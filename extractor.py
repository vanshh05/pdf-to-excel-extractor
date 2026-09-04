"""
Universal PRA Rulebook PDF -> Excel extractor.

    python extractor.py input.pdf -o output.xlsx
    python extractor.py *.pdf -o combined.xlsx --sheet-per-file
"""

import argparse
import os
import sys

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import COLUMNS, STYLE, SPLIT_WORD_THRESHOLD
from textlayer import document_lines, cover_metadata, contents_map
from structure import parse
from emitter import build_rows


def extract(pdf_path):
    meta = cover_metadata(pdf_path)
    lines = list(document_lines(pdf_path, skip_cover=True))
    tree = parse(lines, regulation_name=meta.get("part", ""),
                 chapter_numbers=contents_map(pdf_path))
    rows = build_rows(tree, meta)
    return rows, meta


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _style_sheet(ws, rows):
    header_fill = PatternFill("solid", start_color=STYLE["header_fill"],
                              end_color=STYLE["header_fill"])
    bau_fill = PatternFill("solid", start_color=STYLE["bau_fill"],
                           end_color=STYLE["bau_fill"])
    hname, hsize, hbold, hcolor = STYLE["header_font"]
    bname, bsize, bbold, bcolor = STYLE["body_font"]
    gname, gsize, gbold, gcolor = STYLE["heading_font"]
    header_font = Font(name=hname, size=hsize, bold=hbold, color=hcolor)
    body_font = Font(name=bname, size=bsize, bold=bbold, color=bcolor)
    heading_font = Font(name=gname, size=gsize, bold=gbold, color=gcolor)

    side = Side(style="thin", color=STYLE["border"])
    border = Border(left=side, right=side, top=side, bottom=side)

    for cell in ws[1]:
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.fill = bau_fill if cell.value == "BAU Line C" else header_fill

    n_rows = len(rows)
    for index, excel_row in enumerate(
            ws.iter_rows(min_row=2, max_row=max(n_rows + 1, 2),
                         min_col=1, max_col=len(COLUMNS))):
        is_heading = index < n_rows and rows[index].get("_bold")
        for cell in excel_row:
            cell.font = heading_font if is_heading else body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for letter, width in STYLE["widths"].items():
        ws.column_dimensions[letter].width = width
    for i in range(len(STYLE["widths"]) + 1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.row_dimensions[1].height = STYLE["header_height"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(n_rows + 1, 2)}"


def _rule_cell(row):
    """Rule cell value, with the sub-heading run set in bold.

    The sub-heading is not a column -- it sits at the top of the Rule cell, so
    only that leading run is bolded rather than the whole cell. openpyxl writes
    this as inline rich text.
    """
    text = row.get("Rule", "") or ""
    subheading = row.get("_subheading", "")
    if not subheading or not text.startswith(subheading):
        return text

    name, size, _, color = STYLE["body_font"]
    bold = InlineFont(rFont=name, sz=size, b=True, color=color)
    plain = InlineFont(rFont=name, sz=size, color=color)
    return CellRichText([
        TextBlock(bold, subheading),
        TextBlock(plain, text[len(subheading):]),
    ])


def write_excel(sheets, output_path):
    """sheets: list of (sheet_name, rows)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    rule_col = COLUMNS.index("Rule") + 1 if "Rule" in COLUMNS else None

    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        ws.append(COLUMNS)
        for row in rows:
            ws.append([row.get(col, "") for col in COLUMNS])
        if rule_col:
            for offset, row in enumerate(rows):
                if row.get("_subheading"):
                    ws.cell(offset + 2, rule_col).value = _rule_cell(row)
        _style_sheet(ws, rows)

    wb.save(output_path)


def _sheet_name(pdf_path):
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    for ch in "[]:*?/\\":
        base = base.replace(ch, " ")
    return base.replace("__", " ").replace("_", " ").strip()[:31]


def main():
    ap = argparse.ArgumentParser(
        description="Extract PRA Rulebook PDFs into structured Excel.")
    ap.add_argument("pdfs", nargs="+", help="input PDF file(s)")
    ap.add_argument("-o", "--output", default=None)
    ap.add_argument("--sheet-per-file", action="store_true",
                    help="one worksheet per PDF instead of one combined sheet")
    ap.add_argument("--threshold", type=int, default=SPLIT_WORD_THRESHOLD,
                    help="word count above which a sibling group is split")
    args = ap.parse_args()

    import config
    config.SPLIT_WORD_THRESHOLD = args.threshold
    import emitter
    emitter.SPLIT_WORD_THRESHOLD = args.threshold

    combined = []
    sheets = []

    for path in args.pdfs:
        if not os.path.exists(path):
            print(f"skip (not found): {path}", file=sys.stderr)
            continue
        rows, meta = extract(path)
        print(f"{os.path.basename(path):58s} {len(rows):5d} rows   "
              f"part={meta.get('part','?')}")
        if args.sheet_per_file:
            sheets.append((_sheet_name(path), rows))
        else:
            combined.extend(rows)

    if not args.sheet_per_file:
        sheets = [("Extracted Rules", combined)]

    if not any(rows for _, rows in sheets):
        print("No rules extracted.", file=sys.stderr)
        return 1

    out = args.output or os.path.splitext(args.pdfs[0])[0] + ".xlsx"
    write_excel(sheets, out)
    print(f"\nWritten: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
